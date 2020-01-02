#!/usr/bin/env python
# -*- coding:utf-8 -*-

# openpyxl教程
# 1. 创建工作簿 

from openpyxl import Workbook
from openpyxl import load_workbook


# 保存的文件路径:
files = "F:\\PythonProject\\python-scripts\\pyxltutorial\\balances.xlsx"

# 创建一个工作簿
wb = Workbook()
# 激活第一个工作表来创建工作簿
ws = wb.active

# 创建新工作表
ws1 = wb.create_sheet("6月adaptor日志")
ws2 = wb.create_sheet("7月adaptor日志",1)

# 重命名工作表
# ws.title = "6月adaptor"

# 更改背景颜色
ws.sheet_properties.tabColor = "1072BA"

# 为工作表提供名称,就可以将名称视为工作簿的一个键.
# ws3 = wb["6月adaptor"]

# 查看工作簿的所有工作表的名称
print(wb.sheetnames)

# 遍历工作表
for sheet in wb:
    print(sheet.title)

# 在单个工作簿中创建工作表的副本
# source = wb.active 
# target = wb.copy_worksheet(source)

# 访问数据
# 访问一个单元格数据,单元格可以直接作为工作表的键访问
c = ws['A4']
# A4单元格不存在,就会创建一个单元格,值可以直接分配.
ws['A4'] = 4

# 使用行和列表示法提供对单元格的访问
d = ws.cell(row=4,column=2,value=10)

# 创建100x100个单元格
# for x in range(1,101):
    # for y in range(1,101):
        # ws.cell(row=1,column=y)


# 访问多个单元
# 使用切片访问单元格范围
cell_range = ws['A1':'C2']

# 行或列的范围
colc = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

# 使用iter_rows()方法
for row in ws.iter_rows(min_row=1,max_col=3,max_row=2):
    for cell in row:
        print(cell)

# 使用iter_cols()方法将返回列
for col in ws.iter_cols(min_row=1,max_col=3,max_row=2):
    for cell in col:
        print(cell)
# 注意: iter_cols()方法不能以只读模式使用.


# 遍历文件的所有行或列,使用rows属性.
ws['C9'] = 'test python'
print(tuple(ws.rows))

# columns列, 不以只读模式提供。
print(tuple(ws.columns))

# 获取工作表中的值. 使用values属性,
# 迭代工作表中的所有行,但只返回单元格值.
for row in ws.values:
    for value in row:
        print(value)


# 保存到文件
wb.save(files)


# 从文件加载
# 使用load_workbook()打开现有工作簿:
wb2 = load_workbook('test.xlsx')
print(wb2.sheetnames)